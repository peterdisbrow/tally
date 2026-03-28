from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
P0_FILL = PatternFill('solid', fgColor='FFD7D7')
P1_FILL = PatternFill('solid', fgColor='FFF2CC')
P2_FILL = PatternFill('solid', fgColor='D9E2F3')
P3_FILL = PatternFill('solid', fgColor='E2EFDA')
BLUE_FONT = Font(name='Arial', color='0000FF', size=10)
BLACK_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUB_FONT = Font(name='Arial', bold=True, size=11, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='1F4E79')

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = BLACK_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if fill:
            cell.fill = fill

def section_row(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal='left')
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = SECTION_FILL

# ===== TAB 1: LAUNCH CHECKLIST =====
ws1 = wb.active
ws1.title = "Launch Checklist"
ws1.sheet_properties.tabColor = "1F4E79"

ws1.cell(row=1, column=1, value="Tally Connect — Launch Checklist").font = TITLE_FONT
ws1.cell(row=2, column=1, value="Updated: March 26, 2026").font = Font(name='Arial', italic=True, size=10, color='666666')

headers = ['#', 'Task', 'Category', 'Priority', 'Owner', 'Status', 'Est. Hours', 'Dependencies', 'Notes']
r = 4
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, len(headers))

tasks = [
    # Critical Blockers
    ("CRITICAL BLOCKERS", None),
    (1, "Verify CI passes on latest main push", "Critical", "P0", "Dev", "Not Started", 1, "-", "CI failed on 1b5f956 and 398a795"),
    (2, "Redeploy relay-server to Railway", "Critical", "P0", "Dev", "Not Started", 1, "CI green", "Build failed Mar 21 — production may be stale"),
    (3, "Fix landing page signup plan selection", "Critical", "P0", "Dev", "Not Started", 3, "-", "Everyone lands on Connect regardless of CTA"),
    (4, "Add prices to billing page in portal", "Critical", "P0", "Dev", "Not Started", 4, "-", "Users can't see what they're paying for"),
    (5, "End-to-end signup → trial → upgrade test", "Critical", "P0", "Dev", "Not Started", 4, "Items 1-4", "Full Stripe flow on staging"),
    (6, "Verify auto-update works with real GitHub Release", "Critical", "P0", "Dev", "Not Started", 3, "CI green", "Create test release, verify download + install"),
    (7, "Apple Developer certificate + notarization", "Critical", "P0", "Dev", "Not Started", 4, "-", "macOS Gatekeeper blocks unsigned apps"),

    # Pre-Launch
    ("PRE-LAUNCH", None),
    (8, "Per-user portal login accounts", "Pre-Launch", "P1", "Dev", "Not Started", 16, "-", "Currently 1 shared login per church"),
    (9, "Email notification fallback for alerts", "Pre-Launch", "P1", "Dev", "Not Started", 12, "-", "Telegram is only channel — single point of failure"),
    (10, "Billing page: show tier prices + feature comparison", "Pre-Launch", "P1", "Dev", "Not Started", 6, "Item 4", "Users need to understand what they're upgrading to"),
    (11, "Portal code splitting / lazy loading", "Pre-Launch", "P1", "Dev", "Not Started", 8, "-", "590KB single-file SPA — slow on mobile"),
    (12, "Forgot password link in Electron app", "Pre-Launch", "P1", "Dev", "Not Started", 2, "-", "No in-app path to password recovery"),
    (13, "Complete Spanish i18n coverage", "Pre-Launch", "P1", "Dev", "Not Started", 8, "-", "Login page, billing modals, server errors still English"),
    (14, "Landing page SEO meta tags audit", "Pre-Launch", "P1", "Marketing", "Not Started", 3, "-", "OG image 404, missing structured data"),
    (15, "Create product demo video (2-3 min)", "Pre-Launch", "P1", "Marketing", "Not Started", 12, "-", "Show setup → monitoring → alert flow"),
    (16, "Write 3 launch blog posts", "Pre-Launch", "P1", "Marketing", "Not Started", 8, "-", "Intro, use case story, comparison"),
    (17, "Set up Google Analytics + conversion tracking", "Pre-Launch", "P1", "Marketing", "Not Started", 3, "-", "Track signup funnel"),
    (18, "Privacy policy + Terms of Service pages", "Pre-Launch", "P1", "Ops", "Not Started", 6, "-", "Required before collecting payments"),
    (19, "Support email + helpdesk setup", "Pre-Launch", "P1", "Ops", "Not Started", 3, "-", "support@tallyconnect.app"),
    (20, "Redis/Upstash for production rate limiting", "Pre-Launch", "P1", "Dev", "Not Started", 4, "-", "In-memory rate limiter won't scale"),
    (21, "Stripe webhook endpoint SSL verification", "Pre-Launch", "P2", "Dev", "Not Started", 2, "-", "Confirm webhook signatures in production"),
    (22, "Onboarding email sequence (5 emails)", "Pre-Launch", "P2", "Marketing", "Not Started", 8, "-", "Welcome → Setup → First service → Value → Upgrade"),
    (23, "ROI Calculator wired to landing page", "Pre-Launch", "P2", "Dev", "Not Started", 2, "-", "Component built but never rendered"),
    (24, "Knowledge base / FAQ docs", "Pre-Launch", "P2", "Marketing", "Not Started", 10, "-", "Setup guides, troubleshooting, billing FAQ"),

    # Launch Day
    ("LAUNCH DAY", None),
    (25, "Publish v1.1.0 GitHub Release (signed DMG + EXE)", "Launch", "P0", "Dev", "Not Started", 2, "Items 6,7", "Triggers auto-update for any beta users"),
    (26, "Deploy relay-server to Railway production", "Launch", "P0", "Dev", "Not Started", 1, "All pre-launch done", "Final production deploy"),
    (27, "Publish landing page updates to Vercel", "Launch", "P0", "Dev", "Not Started", 1, "-", "Final copy, pricing, CTA fixes"),
    (28, "Send launch announcement email", "Launch", "P0", "Marketing", "Not Started", 2, "-", "To beta users + waitlist"),
    (29, "Post on church AV Facebook groups", "Launch", "P1", "Marketing", "Not Started", 2, "-", "Church Tech Leaders, AVL, etc."),
    (30, "Post on Reddit r/churchav r/livesound", "Launch", "P1", "Marketing", "Not Started", 1, "-", "Authentic intro, not spammy"),
    (31, "Monitor error logs + Stripe dashboard", "Launch", "P0", "Dev", "Not Started", 4, "-", "All day — watch for signup failures"),

    # Post-Launch
    ("POST-LAUNCH (Weeks 1-4)", None),
    (32, "Daily error log review + hotfix cycle", "Post-Launch", "P0", "Dev", "Not Started", 2, "-", "Per day — first 2 weeks"),
    (33, "Collect NPS from first 10 churches", "Post-Launch", "P1", "Marketing", "Not Started", 3, "-", "Simple survey after 1 week"),
    (34, "Fix top 3 user-reported issues", "Post-Launch", "P1", "Dev", "Not Started", 12, "-", "Budget 4 hrs each"),
    (35, "Church spotlight / case study", "Post-Launch", "P2", "Marketing", "Not Started", 6, "-", "Interview a happy beta church"),
    (36, "Planning Center integration docs", "Post-Launch", "P2", "Dev", "Not Started", 4, "-", "Self-serve setup guide"),
    (37, "PWA manifest for mobile portal", "Post-Launch", "P2", "Dev", "Not Started", 4, "-", "Add to Home Screen support"),
    (38, "Cross-campus unified dashboard", "Post-Launch", "P2", "Dev", "Not Started", 16, "-", "Multi-campus admin single view"),
    (39, "Referral program setup", "Post-Launch", "P3", "Marketing", "Not Started", 6, "-", "Churches referring churches"),
    (40, "Monthly product update email", "Post-Launch", "P3", "Marketing", "Not Started", 3, "-", "Changelog + tips newsletter"),
    (41, "App Store / Homebrew distribution", "Post-Launch", "P3", "Dev", "Not Started", 8, "-", "Easier discovery and install"),
]

r = 5
priority_fills = {'P0': P0_FILL, 'P1': P1_FILL, 'P2': P2_FILL, 'P3': P3_FILL}
for t in tasks:
    if t[1] is None:
        section_row(ws1, r, len(headers), t[0])
        r += 1
        continue
    for i, v in enumerate(t):
        ws1.cell(row=r, column=i+1, value=v)
    fill = priority_fills.get(t[3])
    style_row(ws1, r, len(headers), fill)
    r += 1

# Total hours formula
r += 1
ws1.cell(row=r, column=6, value="TOTAL HOURS:").font = BOLD_FONT
ws1.cell(row=r, column=7).font = BOLD_FONT
ws1.cell(row=r, column=7, value=f'=SUM(G5:G{r-2})')

ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 18
ws1.column_dimensions['I'].width = 40
ws1.auto_filter.ref = f"A4:I{r-2}"

# ===== TAB 2: TIMELINE =====
ws2 = wb.create_sheet("Timeline")
ws2.sheet_properties.tabColor = "2E75B6"

ws2.cell(row=1, column=1, value="Tally Connect — Launch Timeline").font = TITLE_FONT
ws2.cell(row=2, column=1, value="Target Launch: Week of May 25, 2026 (~9 weeks)").font = Font(name='Arial', italic=True, size=10, color='666666')

t_headers = ['Week', 'Dates', 'Phase', 'Dev Tasks (~20 hrs)', 'Marketing/Ops Tasks (~10 hrs)', 'Milestone']
r = 4
for i, h in enumerate(t_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, len(t_headers))

timeline = [
    ("Week 1", "Mar 30 – Apr 5", "Fix Blockers", "CI fixes, Railway redeploy, signup plan selection fix, billing page prices", "Set up analytics, start demo video script", "CI green, production live"),
    ("Week 2", "Apr 6 – 12", "Fix Blockers", "E2E signup→trial→upgrade test, auto-update verification, Apple cert submission", "Draft launch blog post #1, SEO audit fixes", "Stripe flow verified end-to-end"),
    ("Week 3", "Apr 13 – 19", "Core Polish", "Per-user portal accounts (part 1), forgot password link, Redis rate limiting", "Privacy policy + ToS drafts, support email setup", "Portal auth redesign started"),
    ("Week 4", "Apr 20 – 26", "Core Polish", "Per-user portal accounts (part 2), email notification fallback", "Onboarding email sequence (5 emails), knowledge base start", "Multi-user portal live"),
    ("Week 5", "Apr 27 – May 3", "Launch Prep", "Portal code splitting, Spanish i18n completion, ROI calculator wiring", "Demo video production, blog posts #2 and #3", "Portal under 200KB initial load"),
    ("Week 6", "May 4 – 10", "Beta Outreach", "Bug fixes from beta feedback, webhook SSL verification", "Reach out to 10 churches for beta, FAQ docs", "5 beta churches actively using"),
    ("Week 7", "May 11 – 17", "Beta Outreach", "Top beta bug fixes, Stripe edge cases", "Collect beta feedback, refine messaging, prep launch email", "10 beta churches, feedback collected"),
    ("Week 8", "May 18 – 24", "Final Prep", "Final QA pass, staging → production, notarized build", "Launch email drafted, social posts queued", "Launch-ready build signed & tested"),
    ("Week 9", "May 25 – 31", "LAUNCH", "Publish release, deploy production, monitor logs all week", "Launch email, Facebook groups, Reddit, Product Hunt?", "🚀 LIVE — first paid signups"),
    ("Week 10", "Jun 1 – 7", "Post-Launch", "Daily hotfix cycle, top 3 reported issues", "NPS survey to early users, monitor conversion funnel", "First MRR milestone"),
    ("Week 11", "Jun 8 – 14", "Post-Launch", "Planning Center integration docs, PWA manifest", "Church spotlight / case study interview", "Case study published"),
    ("Week 12", "Jun 15 – 21", "Post-Launch", "Cross-campus dashboard (part 1)", "Referral program setup, monthly newsletter", "Multi-campus feature preview"),
    ("Week 13", "Jun 22 – 28", "Post-Launch", "Cross-campus dashboard (part 2), App Store prep", "Product update email, analyze churn", "30-day post-launch review"),
]

LAUNCH_FILL = PatternFill('solid', fgColor='E2EFDA')
for i, t in enumerate(timeline):
    r = 5 + i
    for j, v in enumerate(t):
        ws2.cell(row=r, column=j+1, value=v)
    fill = LAUNCH_FILL if "LAUNCH" in t[2] else None
    style_row(ws2, r, len(t_headers), fill)

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 55
ws2.column_dimensions['E'].width = 45
ws2.column_dimensions['F'].width = 35

# ===== TAB 3: GOALS & METRICS =====
ws3 = wb.create_sheet("Goals & Metrics")
ws3.sheet_properties.tabColor = "548235"

ws3.cell(row=1, column=1, value="Tally Connect — Goals & Metrics").font = TITLE_FONT

g_headers = ['Metric', 'Target', 'Stretch Goal', 'How to Measure', 'Notes']
r = 3
for i, h in enumerate(g_headers, 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, len(g_headers))

goals = [
    ("PRE-LAUNCH (Weeks 1-8)", None),
    ("Beta churches onboarded", 10, 15, "Stripe customer count (trial)", "Personal outreach to local churches + online communities"),
    ("Open bugs (P0/P1)", 0, 0, "GitHub Issues", "Zero P0/P1 at launch gate"),
    ("CI pass rate", "100%", "100%", "GitHub Actions", "All 3 test suites green"),
    ("Lighthouse score (landing page)", "90+", "95+", "PageSpeed Insights", "Performance, SEO, Accessibility"),
    ("Beta NPS score", "40+", "60+", "Survey (1-10 scale)", "Anything above 30 is good for beta"),

    ("LAUNCH WEEK (Week 9)", None),
    ("New signups", 25, 50, "Stripe + DB", "Free + paid combined"),
    ("Paid conversions (trial → Plus/Pro)", 3, 8, "Stripe dashboard", "~12% conversion rate target"),
    ("Landing page → signup conversion", "5%", "8%", "Google Analytics funnel", "Industry avg is 2-5% for SaaS"),
    ("Uptime", "99.9%", "100%", "Health endpoint monitor", "No downtime during launch week"),
    ("Support tickets", "<10", "<5", "Email inbox", "Low = good onboarding"),

    ("30-DAY (Week 13)", None),
    ("Total churches (free + paid)", 40, 75, "DB count", "Includes free Connect tier"),
    ("Paying churches", 8, 15, "Stripe active subscriptions", "Plus, Pro, or Enterprise"),
    ("MRR", "$350", "$700", "Stripe MRR dashboard", "8 × avg $44/mo"),
    ("Monthly churn rate", "<5%", "<3%", "Stripe cancellations / active", "First-month churn is typically higher"),
    ("NPS score", "50+", "65+", "Survey", "Good = 50+, Great = 70+"),

    ("90-DAY (Month 3)", None),
    ("Total churches", 80, 150, "DB count", "Organic + referral growth"),
    ("Paying churches", 20, 40, "Stripe", "25% paid conversion rate"),
    ("MRR", "$1,200", "$2,500", "Stripe", "Blended avg ~$60/mo"),
    ("Annual run rate", "$14,400", "$30,000", "MRR × 12", "ARR milestone"),
    ("Referral signups", 5, 15, "Referral tracking", "Churches referring churches"),
    ("Feature requests logged", 20, 30, "GitHub Issues / feedback", "Signal of engagement"),
]

r = 4
for g in goals:
    if g[1] is None:
        section_row(ws3, r, len(g_headers), g[0])
        r += 1
        continue
    for i, v in enumerate(g):
        ws3.cell(row=r, column=i+1, value=v)
    style_row(ws3, r, len(g_headers))
    r += 1

ws3.column_dimensions['A'].width = 38
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 45

# ===== TAB 4: REVENUE PROJECTIONS =====
ws4 = wb.create_sheet("Revenue Projections")
ws4.sheet_properties.tabColor = "BF8F00"

ws4.cell(row=1, column=1, value="Tally Connect — Revenue Projections (6-Month)").font = TITLE_FONT

# Pricing table
ws4.cell(row=3, column=1, value="Pricing Table").font = SUB_FONT
price_headers = ['Tier', 'Monthly Price']
for i, h in enumerate(price_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header(ws4, 4, 2)

prices = [("Connect", 0), ("Plus", 29), ("Pro", 79), ("Enterprise", 199)]
for i, (tier, price) in enumerate(prices):
    r = 5 + i
    ws4.cell(row=r, column=1, value=tier).font = BLACK_FONT
    ws4.cell(row=r, column=2, value=price).font = BLUE_FONT
    ws4.cell(row=r, column=2).number_format = '$#,##0'
    for c in range(1, 3):
        ws4.cell(row=r, column=c).border = THIN_BORDER

# Conservative scenario
scenarios = [
    ("Conservative", 12, [
        ("Connect", [5, 10, 15, 20, 25, 30]),
        ("Plus", [1, 2, 4, 6, 8, 10]),
        ("Pro", [0, 1, 1, 2, 3, 4]),
        ("Enterprise", [0, 0, 0, 0, 0, 1]),
    ]),
    ("Moderate", 22, [
        ("Connect", [8, 18, 30, 40, 50, 60]),
        ("Plus", [2, 5, 8, 12, 16, 20]),
        ("Pro", [1, 2, 4, 6, 8, 10]),
        ("Enterprise", [0, 0, 1, 1, 2, 2]),
    ]),
    ("Optimistic", 32, [
        ("Connect", [15, 30, 50, 70, 90, 110]),
        ("Plus", [3, 8, 15, 22, 30, 38]),
        ("Pro", [2, 4, 8, 12, 16, 20]),
        ("Enterprise", [0, 1, 2, 3, 4, 5]),
    ]),
]

months = ["Month 1\nJun '26", "Month 2\nJul '26", "Month 3\nAug '26", "Month 4\nSep '26", "Month 5\nOct '26", "Month 6\nNov '26"]

current_row = 11
summary_refs = {}

for scenario_name, _, tiers in scenarios:
    ws4.cell(row=current_row, column=1, value=f"{scenario_name} Scenario").font = SUB_FONT
    current_row += 1

    # Sub headers
    sub_h = ['', 'Subscribers'] + months
    for i, h in enumerate(sub_h, 1):
        ws4.cell(row=current_row, column=i, value=h)
    style_header(ws4, current_row, len(sub_h))
    current_row += 1

    tier_start = current_row
    for tier_name, counts in tiers:
        ws4.cell(row=current_row, column=1, value=tier_name).font = BLACK_FONT
        ws4.cell(row=current_row, column=1).border = THIN_BORDER
        ws4.cell(row=current_row, column=2, value="Count").font = Font(name='Arial', italic=True, size=9, color='666666')
        ws4.cell(row=current_row, column=2).border = THIN_BORDER
        for j, count in enumerate(counts):
            cell = ws4.cell(row=current_row, column=3+j, value=count)
            cell.font = BLUE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

    # MRR row per tier
    ws4.cell(row=current_row, column=1, value="").border = THIN_BORDER
    ws4.cell(row=current_row, column=2, value="MRR").font = BOLD_FONT
    ws4.cell(row=current_row, column=2).border = THIN_BORDER

    # Price refs: B5=Connect($0), B6=Plus($29), B7=Pro($79), B8=Enterprise($199)
    for j in range(6):
        col = get_column_letter(3+j)
        tier_rows = [tier_start + k for k in range(4)]
        # MRR = Connect*$0 + Plus*$29 + Pro*$79 + Enterprise*$199
        formula = f'={col}{tier_rows[0]}*$B$5+{col}{tier_rows[1]}*$B$6+{col}{tier_rows[2]}*$B$7+{col}{tier_rows[3]}*$B$8'
        cell = ws4.cell(row=current_row, column=3+j, value=formula)
        cell.font = BOLD_FONT
        cell.number_format = '$#,##0'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    mrr_row = current_row
    summary_refs[scenario_name] = mrr_row

    # Total subscribers row
    current_row += 1
    ws4.cell(row=current_row, column=1, value="").border = THIN_BORDER
    ws4.cell(row=current_row, column=2, value="Total Subs").font = BOLD_FONT
    ws4.cell(row=current_row, column=2).border = THIN_BORDER
    for j in range(6):
        col = get_column_letter(3+j)
        formula = f'=SUM({col}{tier_start}:{col}{tier_start+3})'
        cell = ws4.cell(row=current_row, column=3+j, value=formula)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    current_row += 2

# Summary comparison
ws4.cell(row=current_row, column=1, value="Scenario Comparison — Month 6 MRR").font = SUB_FONT
current_row += 1
comp_headers = ['Scenario', 'Month 6 MRR', 'Month 6 ARR', 'Month 6 Paid Subs']
for i, h in enumerate(comp_headers, 1):
    ws4.cell(row=current_row, column=i, value=h)
style_header(ws4, current_row, len(comp_headers))
current_row += 1

for scenario_name in ['Conservative', 'Moderate', 'Optimistic']:
    mrr_r = summary_refs[scenario_name]
    ws4.cell(row=current_row, column=1, value=scenario_name).font = BLACK_FONT
    ws4.cell(row=current_row, column=1).border = THIN_BORDER

    mrr_cell = ws4.cell(row=current_row, column=2, value=f'=H{mrr_r}')
    mrr_cell.font = BOLD_FONT
    mrr_cell.number_format = '$#,##0'
    mrr_cell.border = THIN_BORDER

    arr_cell = ws4.cell(row=current_row, column=3, value=f'=B{current_row}*12')
    arr_cell.font = BOLD_FONT
    arr_cell.number_format = '$#,##0'
    arr_cell.border = THIN_BORDER

    # Paid subs = Plus + Pro + Enterprise for month 6 (col H)
    tier_start_for = summary_refs[scenario_name] - 1  # total subs row is after MRR
    # Actually need to find the subscriber rows. MRR row = summary_refs[scenario_name]
    # Subscriber rows are MRR - 4 through MRR - 1
    sub_start = mrr_r - 4
    paid_formula = f'=H{sub_start+1}+H{sub_start+2}+H{sub_start+3}'
    paid_cell = ws4.cell(row=current_row, column=4, value=paid_formula)
    paid_cell.font = BOLD_FONT
    paid_cell.border = THIN_BORDER

    current_row += 1

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 16
for c in range(3, 9):
    ws4.column_dimensions[get_column_letter(c)].width = 14

# Freeze panes
ws1.freeze_panes = 'A5'
ws2.freeze_panes = 'A5'
ws3.freeze_panes = 'A4'

output = '/sessions/fervent-sleepy-pasteur/mnt/church-av/Tally_Connect_Launch_Plan.xlsx'
wb.save(output)
print(f"Saved to {output}")
